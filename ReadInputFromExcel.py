import math

import openpyxl

import Util
from Path import Path
from Station import Station
from Truck import Truck

stations = []
paths = []
trucks = []
NUM_OF_STATIONS_WITH_DEPOT = 0


# Read number of stations with depot from the input file
def extract_number_of_stations():
    global NUM_OF_STATIONS_WITH_DEPOT
    file_location = "input.xlsx"
    work_book = openpyxl.load_workbook(file_location)
    timing_sheet = work_book.worksheets[5]
    NUM_OF_STATIONS_WITH_DEPOT = int(timing_sheet.cell(2, 2).value) + 1


# Read demands of stations
def extract_station_demand_sheet():
    file_location = "input.xlsx"
    work_book = openpyxl.load_workbook(file_location)
    station_demand_sheet = work_book.worksheets[2]
    stations.append(Station(0, float("inf")))  # The first station is always dummy, used as start-end point.
    for index in range(1, NUM_OF_STATIONS_WITH_DEPOT):
        stations.append(Station(index, int(math.ceil(station_demand_sheet.cell(index, 1).value))))
    extract_time_constraints()
    return stations


# Read earliest, latest convenience time for the stations and average time spent for the service
def extract_time_constraints():
    file_location = "input.xlsx"
    work_book = openpyxl.load_workbook(file_location)
    timing_sheet = work_book.worksheets[4]
    for index in range(2, NUM_OF_STATIONS_WITH_DEPOT + 2):
        stations[index - 2].earliest_time = timing_sheet.cell(index, 1).value
        stations[index - 2].latest_time = timing_sheet.cell(index, 2).value
        stations[index - 2].time_to_spend = timing_sheet.cell(index, 3).value


# Populate Paths by reading each cell of the distance matrix. Savings will be calculated later. Not ready in input file.
def extract_distance_sheet() -> []:
    file_location = "input.xlsx"
    work_book = openpyxl.load_workbook(file_location)
    station_distance_sheet = work_book.worksheets[3]
    time_between_stations_sheet = work_book.worksheets[1]

    for row in range(1, NUM_OF_STATIONS_WITH_DEPOT + 1):
        for column in range(row + 1, NUM_OF_STATIONS_WITH_DEPOT + 1):
            paths.append(Path(stations[row - 1], stations[column - 1],
                              int(math.ceil(station_distance_sheet.cell(row, column).value)), 0.0,
                              time_between_stations_sheet.cell(row, column).value))
    return paths


# Read trucks from the input file
def extract_truck_sheet():
    file_location = "input.xlsx"
    work_book = openpyxl.load_workbook(file_location)
    truck_sheet = work_book.worksheets[0]
    for row in range(2, 4):
        trucks.append(
            Truck(truck_sheet.cell(row, 1).value, truck_sheet.cell(row, 2).value, truck_sheet.cell(row, 3).value))
    return trucks


# Save results in the Excel file after algorithm converges
def print_results_into_excel(routes):
    file_location = "input.xlsx"
    work_book = openpyxl.load_workbook(file_location)
    result_sheet = work_book.worksheets[6]
    result_sheet.cell(1, 1).value = "ROUTE NO"
    result_sheet.cell(1, 2).value = "TOTAL CARRIED DEMAND"
    result_sheet.cell(1, 3).value = "TOTAL COST OF ROUTE"
    result_sheet.cell(1, 4).value = "ROUTE DURATION"
    result_sheet.cell(1, 5).value = "TRUCK TYPE"
    result_sheet.cell(1, 6).value = "STATIONS"
    for index, route in enumerate(routes):
        result_sheet.cell(index + 2, 1).value = index + 1
        result_sheet.cell(index + 2, 2).value = Util.total_demand_of_route(route)
        result_sheet.cell(index + 2, 3).value = Util.calculate_cost_of_route(route,
                                                                             Util.total_demand_of_route(route) > trucks[
                                                                              0].capacity)
        result_sheet.cell(index + 2, 4).value = Util.calculate_final_time(route)
        if Util.total_demand_of_route(route) > trucks[0].capacity:
            result_sheet.cell(index + 2, 5).value = "TYPE 2"
        else:
            result_sheet.cell(index + 2, 5).value = "TYPE 1"
        result_sheet.cell(index + 2, 6).value = 0
        for i, station in enumerate(route.inner_stations):
            result_sheet.cell(index + 2, i + 7).value = station.stationNum
        result_sheet.cell(index + 2, len(route.inner_stations) + 7).value = 0
    work_book.save(file_location)
